class Data:
          
    def get_data_by_sheet_name(self, sFilePath, sSheetName):
    #"""Get data from given file path, sheet name then return array of object""" 
        from openpyxl import load_workbook
        from openpyxl import Workbook

        wb = load_workbook(sFilePath.strip())
        ws = wb.get_sheet_by_name(sSheetName)
        row_count = ws.get_highest_row()
        column_count = ws.get_highest_column()
        #init data returned   
        test_data = []
    
        for i in range(2, row_count + 1):
            #ini dict
            row_data = {}
            for j in range(1, column_count + 1):
                if str(ws.cell(row = 1, column = j).value) == 'None':
                    continue
                else:
                    key = ws.cell(row = 1, column = j).value                 
                    value =  ws.cell(row = i, column = j).value
                    if value == None:
                        value = ''
                    row_data[key] = str(value)
            #print row_data
            test_data.append(row_data.copy())

        return test_data
    #end get_data_by_sheet_name 



##a = Data()
##b = a.get_data_by_sheet_name('D:\\Workplace\\FS\\fs-qc\\WEB_AT\\DataSet\\hihi.xlsx','Sheet1')
##print b
##raw_input('Enter to exit...')
